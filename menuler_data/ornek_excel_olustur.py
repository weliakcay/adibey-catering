#!/usr/bin/env python3
"""
Bu script, menuler_data/ klasörüne örnek Excel dosyaları oluşturur.
Gerçek menü verilerinizi girmek için bu dosyaları düzenleyin.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TURUNCU = "FFEC5B13"
KOYU    = "FF111111"
BEYAZ   = "FFFFFFFF"
GRI     = "FF2a2a2a"

def stil_baslik(ws, row, col, deger):
    c = ws.cell(row=row, column=col, value=deger)
    c.font = Font(bold=True, color=BEYAZ, size=12)
    c.fill = PatternFill("solid", fgColor=TURUNCU)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def stil_veri(ws, row, col, deger):
    c = ws.cell(row=row, column=col, value=deger)
    c.font = Font(color=BEYAZ, size=11)
    c.fill = PatternFill("solid", fgColor=GRI if row % 2 == 0 else KOYU)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    return c

def ayarla_sutunlar(ws, genislikler):
    for harf, gen in genislikler.items():
        ws.column_dimensions[harf].width = gen

# ─────────────────────────────────────────────
# 1. TABLDOT MENÜLERİ
# ─────────────────────────────────────────────
def olustur_tabldot():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabldot Menüleri"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    basliklar = ["Menü Adı", "Etiket", "Etiket Rengi", "Fiyat (₺/kişi)", "Yemek 1", "Yemek 2", "Yemek 3", "Yemek 4", "Yemek 5"]
    for i, b in enumerate(basliklar, 1):
        stil_baslik(ws, 1, i, b)

    # Örnek veriler — bunları kendi menülerinizle değiştirin
    satirlar = [
        ["Anadolu Esintisi", "Haftalık Favori", "bg-primary",     170, "Tavuk Çorbası",        "Makarna",            "Tavuk Sarma",        "",               ""],
        ["Fit & Form",       "Düşük Kalori",    "bg-emerald-600", 160, "Brokoli Kremalı Çorba","Izgara Tavuk Göğsü", "Kinoa Salatası",     "Şekersiz Elma Tatlısı", ""],
        ["Usta İşi Lezzetler","",               "",               155, "Yayla Çorbası",        "Kadınbudu Köfte",    "Patates Püresi",     "Fırın Sütlaç",   ""],
    ]

    for r, satir in enumerate(satirlar, 2):
        ws.row_dimensions[r].height = 22
        for c, deger in enumerate(satir, 1):
            stil_veri(ws, r, c, deger)

    ayarla_sutunlar(ws, {"A":22,"B":18,"C":18,"D":14,"E":22,"F":22,"G":22,"H":22,"I":22})

    yol = os.path.join(BASE_DIR, "tabldot_menuler.xlsx")
    wb.save(yol)
    print(f"✅ {yol}")

# ─────────────────────────────────────────────
# 2. KOKTEYL MENÜLERİ
# ─────────────────────────────────────────────
def olustur_kokteyl():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kokteyl Menüleri"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    basliklar = ["Menü Adı", "Alt Başlık", "Ürün 1", "Ürün 2", "Ürün 3", "Ürün 4", "Ürün 5"]
    for i, b in enumerate(basliklar, 1):
        stil_baslik(ws, 1, i, b)

    satirlar = [
        ["Modern Kokteyl", "Finger food çeşitleri ve mini sunumlarla şık resepsiyonlar.", "Karides Tempura", "Somon Tartar", "Roast Beef Roll", "Bruschetta", ""],
    ]

    for r, satir in enumerate(satirlar, 2):
        ws.row_dimensions[r].height = 22
        for c, deger in enumerate(satir, 1):
            stil_veri(ws, r, c, deger)

    ayarla_sutunlar(ws, {"A":22,"B":45,"C":20,"D":20,"E":20,"F":20,"G":20})

    yol = os.path.join(BASE_DIR, "kokteyl_menuler.xlsx")
    wb.save(yol)
    print(f"✅ {yol}")

# ─────────────────────────────────────────────
# 3. VIP MENÜLERİ
# ─────────────────────────────────────────────
def olustur_vip():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VIP Menüleri"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    basliklar = ["Menü Adı", "Alt Başlık", "Ürün 1", "Ürün 2", "Ürün 3", "Ürün 4", "Ürün 5"]
    for i, b in enumerate(basliklar, 1):
        stil_baslik(ws, 1, i, b)

    satirlar = [
        ["VIP Gala Dinner", "Siyah kravat etkinlikleri için 5 aşamalı fine-dining menüler.", "Antrikot Confite", "Truffle Risotto", "Altın Yapraklı Sufle", "", ""],
    ]

    for r, satir in enumerate(satirlar, 2):
        ws.row_dimensions[r].height = 22
        for c, deger in enumerate(satir, 1):
            stil_veri(ws, r, c, deger)

    ayarla_sutunlar(ws, {"A":22,"B":50,"C":22,"D":22,"E":22,"F":22,"G":22})

    yol = os.path.join(BASE_DIR, "vip_menuler.xlsx")
    wb.save(yol)
    print(f"✅ {yol}")

if __name__ == "__main__":
    olustur_tabldot()
    olustur_kokteyl()
    olustur_vip()
    print("\nÖrnek Excel dosyaları oluşturuldu.")
    print("Verileri doldurup 'menuler_guncelle.py' scriptini çalıştırın.")
