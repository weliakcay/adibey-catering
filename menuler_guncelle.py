#!/usr/bin/env python3
"""
Adıbey Catering — Menü Güncelleme Scripti
==========================================
Kullanım:
    python3 menuler_guncelle.py

Excel dosyalarını okur ve menuler.html sayfasını otomatik günceller.
Excel dosyaları: menuler_data/ klasöründe
"""

import openpyxl
import os
import re
import shutil
from datetime import datetime

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DATA_DIR   = os.path.join(BASE_DIR, "menuler_data")
HTML_DOSYA = os.path.join(BASE_DIR, "menuler.html")

# ─────────────────────────────────────────────
# EXCEL OKUYUCULAR
# ─────────────────────────────────────────────

def excel_oku(dosya_adi):
    yol = os.path.join(DATA_DIR, dosya_adi)
    if not os.path.exists(yol):
        print(f"⚠️  Bulunamadı: {dosya_adi} — atlandı.")
        return []
    wb = openpyxl.load_workbook(yol)
    ws = wb.active
    satirlar = []
    basliklar = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        satirlar.append(dict(zip(basliklar, row)))
    return satirlar


def deger(satir, *anahtarlar, varsayilan=""):
    for k in anahtarlar:
        v = satir.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return varsayilan


# ─────────────────────────────────────────────
# HTML OLUŞTURUCULAR
# ─────────────────────────────────────────────

def tabldot_kart_html(satir, index):
    ad       = deger(satir, "Menü Adı", varsayilan="Menü")
    etiket   = deger(satir, "Etiket")
    et_renk  = deger(satir, "Etiket Rengi", varsayilan="bg-primary")
    fiyat    = deger(satir, "Fiyat (₺/kişi)", "Fiyat", varsayilan="—")
    yemekler = [deger(satir, f"Yemek {i}") for i in range(1, 8)]
    yemekler = [y for y in yemekler if y]

    # Görsel: daha sonra eklenecek — şimdilik placeholder
    gorsel_html = f"""
                    <!-- GÖRSEL: images/menu_{index+1}.jpg dosyasını bu klasöre ekleyin -->
                    <div class="w-full h-full bg-navy-muted flex items-center justify-center">
                        <div class="text-center text-white/30">
                            <i class="fa-solid fa-image text-4xl mb-2"></i>
                            <p class="text-xs">Görsel eklenecek</p>
                        </div>
                    </div>"""

    etiket_html = ""
    if etiket:
        etiket_html = f"""
                        <div class="absolute top-4 left-4">
                            <span class="{et_renk} text-white text-[10px] font-bold px-3 py-1 rounded-full uppercase">{etiket}</span>
                        </div>"""

    yemek_listeleri = "\n".join(
        f'                            <li class="flex items-center gap-2">{y}</li>'
        for y in yemekler
    )

    return f"""                <!-- {ad} -->
                <div class="group bg-[#111111] rounded-2xl overflow-hidden shadow-2xl transition-all border border-white/5 hover:border-primary/50">
                    <div class="relative h-64 overflow-hidden">{gorsel_html}{etiket_html}
                    </div>
                    <div class="p-8">
                        <div class="flex justify-between items-start mb-4">
                            <h3 class="text-xl font-bold text-white">{ad}</h3>
                            <span class="text-primary font-bold">{fiyat}₺ <small class="text-[10px] text-slate-400">/kişi</small></span>
                        </div>
                        <ul class="space-y-3 mb-8 text-sm text-white/60">
{yemek_listeleri}
                        </ul>
                        <a href="teklif-al.html"
                            class="block w-full py-3 rounded-lg border border-white/10 text-white font-bold hover:bg-primary hover:border-primary transition-all text-center">Detaylı İncele</a>
                    </div>
                </div>"""


def kokteyl_kart_html(satir):
    ad        = deger(satir, "Menü Adı", varsayilan="Kokteyl")
    alt_baslik= deger(satir, "Alt Başlık", varsayilan="")
    urunler   = [deger(satir, f"Ürün {i}") for i in range(1, 8)]
    urunler   = [u for u in urunler if u]

    badge_html = "\n".join(
        f'                            <span class="bg-white/10 backdrop-blur-md px-4 py-2 rounded-full text-xs font-semibold text-white">{u}</span>'
        for u in urunler
    )

    return f"""                <!-- {ad} -->
                <div class="relative group h-[500px] rounded-3xl overflow-hidden shadow-2xl">
                    <!-- GÖRSEL: images/kokteyl_1.jpg eklenecek -->
                    <div class="w-full h-full bg-navy-muted flex items-center justify-center">
                        <div class="text-center text-white/20">
                            <i class="fa-solid fa-champagne-glasses text-6xl mb-2"></i>
                            <p class="text-sm">Görsel eklenecek</p>
                        </div>
                    </div>
                    <div class="absolute inset-0 bg-gradient-to-t from-black via-black/20 to-transparent"></div>
                    <div class="absolute bottom-0 left-0 p-10 w-full">
                        <h3 class="text-3xl font-extrabold text-white mb-2">{ad}</h3>
                        <p class="text-slate-300 mb-6">{alt_baslik}</p>
                        <div class="flex flex-wrap gap-2 mb-8">
{badge_html}
                        </div>
                        <a href="teklif-al.html"
                            class="inline-block bg-primary text-white px-8 py-3 rounded-xl font-bold hover:translate-x-2 transition-transform">Kataloğu Görüntüle</a>
                    </div>
                </div>"""


def vip_kart_html(satir):
    ad        = deger(satir, "Menü Adı", varsayilan="VIP")
    alt_baslik= deger(satir, "Alt Başlık", varsayilan="")
    urunler   = [deger(satir, f"Ürün {i}") for i in range(1, 8)]
    urunler   = [u for u in urunler if u]

    badge_html = "\n".join(
        f'                            <span class="bg-white/10 backdrop-blur-md px-4 py-2 rounded-full text-xs font-semibold text-white">{u}</span>'
        for u in urunler
    )

    return f"""                <!-- {ad} -->
                <div class="relative group h-[500px] rounded-3xl overflow-hidden shadow-2xl">
                    <!-- GÖRSEL: images/vip_1.jpg eklenecek -->
                    <div class="w-full h-full bg-navy-muted flex items-center justify-center">
                        <div class="text-center text-white/20">
                            <i class="fa-solid fa-star text-6xl mb-2"></i>
                            <p class="text-sm">Görsel eklenecek</p>
                        </div>
                    </div>
                    <div class="absolute inset-0 bg-gradient-to-t from-black via-black/20 to-transparent"></div>
                    <div class="absolute bottom-0 left-0 p-10 w-full">
                        <h3 class="text-3xl font-extrabold text-white mb-2">{ad}</h3>
                        <p class="text-slate-300 mb-6">{alt_baslik}</p>
                        <div class="flex flex-wrap gap-2 mb-8">
{badge_html}
                        </div>
                        <a href="teklif-al.html"
                            class="inline-block bg-primary text-white px-8 py-3 rounded-xl font-bold hover:translate-x-2 transition-transform">Özel Teklif Al</a>
                    </div>
                </div>"""


# ─────────────────────────────────────────────
# HTML GÜNCELLE
# ─────────────────────────────────────────────

def html_guncelle(tabldot_html, kokteyl_html, vip_html):
    with open(HTML_DOSYA, "r", encoding="utf-8") as f:
        icerik = f.read()

    # Yedek al
    yedek = HTML_DOSYA.replace(".html", f"_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
    shutil.copy(HTML_DOSYA, yedek)
    print(f"💾 Yedek alındı: {os.path.basename(yedek)}")

    # TABLDOT bölümünü değiştir
    tabldot_blok = f"""            <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-8">
{tabldot_html}
            </div>"""
    icerik = re.sub(
        r'(<div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-8">)'
        r'.*?'
        r'(</div>\s*</div>\s*<!-- Chef)',
        tabldot_blok + "\n        </div>\n\n        <!-- Chef",
        icerik,
        flags=re.DOTALL,
        count=1
    )

    # KOKTEYL + VIP bölümünü değiştir
    kokteylvip_blok = f"""            <div class="grid grid-cols-1 lg:grid-cols-2 gap-12">
{kokteyl_html}
{vip_html}
            </div>"""
    icerik = re.sub(
        r'(<div class="grid grid-cols-1 lg:grid-cols-2 gap-12">)'
        r'.*?'
        r'(</div>\s*</div>\s*</main>)',
        kokteylvip_blok + "\n        </div>\n    </main>",
        icerik,
        flags=re.DOTALL,
        count=1
    )

    with open(HTML_DOSYA, "w", encoding="utf-8") as f:
        f.write(icerik)

    print(f"✅ menuler.html güncellendi!")


# ─────────────────────────────────────────────
# ANA AKIŞ
# ─────────────────────────────────────────────

def main():
    print("=" * 50)
    print("  Adıbey Catering — Menü Güncelleme")
    print("=" * 50)

    # Tabldot
    tabldot_satirlar = excel_oku("tabldot_menuler.xlsx")
    if tabldot_satirlar:
        tabldot_kartlar = "\n".join(tabldot_kart_html(s, i) for i, s in enumerate(tabldot_satirlar))
        print(f"📋 Tabldot menüler okundu: {len(tabldot_satirlar)} adet")
    else:
        tabldot_kartlar = "                <!-- Tabldot menü bulunamadı -->"

    # Kokteyl
    kokteyl_satirlar = excel_oku("kokteyl_menuler.xlsx")
    if kokteyl_satirlar:
        kokteyl_kartlar = "\n".join(kokteyl_kart_html(s) for s in kokteyl_satirlar)
        print(f"🥂 Kokteyl menüler okundu: {len(kokteyl_satirlar)} adet")
    else:
        kokteyl_kartlar = "                <!-- Kokteyl menü bulunamadı -->"

    # VIP
    vip_satirlar = excel_oku("vip_menuler.xlsx")
    if vip_satirlar:
        vip_kartlar = "\n".join(vip_kart_html(s) for s in vip_satirlar)
        print(f"⭐ VIP menüler okundu: {len(vip_satirlar)} adet")
    else:
        vip_kartlar = "                <!-- VIP menü bulunamadı -->"

    html_guncelle(tabldot_kartlar, kokteyl_kartlar, vip_kartlar)
    print("=" * 50)
    print("Tamamlandı! Sayfayı tarayıcıda yenileyin.")
    print("=" * 50)


if __name__ == "__main__":
    main()
